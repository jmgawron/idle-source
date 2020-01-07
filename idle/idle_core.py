from openpyxl import load_workbook
import os, jinja2, json, base64, getpass, datetime, time, urllib3
from settings.dnac_settings import BASE_URI, USER, PASSWORD
import requests
from idle import add_ip_pool, add_virtual_network, add_sg_to_vn, add_site_group, add_servers, add_ip_pool_to_vn
from idle import add_ise_aaa, set_aaa_properties, add_cli, make_primary_creds, add_snmpv2c, add_snmpv3
from idle import add_network_settings, add_wireless_profile, add_ent_wlan, add_fabric, set_auth_template
from urllib3.exceptions import InsecureRequestWarning

urllib3.disable_warnings(InsecureRequestWarning)


# Static values
CWD = os.path.abspath(__file__)[:-len(__file__)]
TEMPLATE_DIR = 'templates' + '/'
SEPARATOR = "=========================================================================================================="


class DNACBuildTask(object):
    #
    # I'm to be a DNACBuildTask object with the following attributes (please note use of kwargs to define attributes):
    #   template:           { I am whatever is defined in the spreadsheet to be pushed to DNAC API }
    #   method:
    #
    def __init__(self, template, task):
        self.template = template
        self.task = task
        self.status = None
        self.reason = None
        self.complete_time = None

    def set_attribute(self, **kwargs):
        for key, value in kwargs.items():
            setattr(self, key, value)

    def print_task(self):
        print('+template={}, complete_time={}, status={}, reason={}'.format(
            self.template, self.complete_time, self.status, self.reason))

    def print_all(self):
        print('+', vars(self))

    def print_jinja(self):
        print('+jinja={}'.format(self.jinja))


class TaskRunner(object):
    def __init__(self, filename):
        self.connector = DNACConnector()
        self.build_tasks = []
        self.filename = filename
        self.inputQueue = []
        self.logger = Logger()

    def gate_keeper(func):
        def wrapper(self):
            while True:
                func(self)
                self.get_wb_data()
                self.print_queued_tasks()
                answer = input("Would you like push the queued tasks or start over? y/n ")
                if answer in ['Y', 'y']:
                    break
                else:
                    exit(0)
        return wrapper

    def process_tasks_debug(self):
        for task in self.inputQueue:
            # Call external module to get any prerequisites needed to populate template and PUT/POST
            if 'add_ip_pool' in task.task:
                add_ip_pool.update_ip_pool_task(task, self.populate_template)
            elif 'add_virtual_network' in task.task:
                add_virtual_network.update_add_virtual_network_task(task, self.populate_template)
            elif 'add_scalable_group_to_vn' in task.task:
                add_sg_to_vn.update_add_sg_to_vn_task(task, self.populate_template, self.connector)
            elif 'add_site_group' in task.task:
                add_site_group.update_add_sg_to_vn_task(task, self.populate_template, self.connector)
            elif 'add_servers' in task.task:
                add_servers.update_add_servers(task, self.populate_template, self.connector)
            elif 'add_ip_pool_to_vn' in task.task:
                add_ip_pool_to_vn.update_add_ip_pool_to_vn_task(task, self.populate_template, self.connector)
            elif 'add_ise_aaa' in task.task:
                add_ise_aaa.update_add_ise_aaa_task(task, self.populate_template)
            elif 'set_aaa_properties' in task.task:
                set_aaa_properties.update_set_aaa_properties_task(task, self.populate_template, self.connector)
            elif 'add_cli' in task.task:
                add_cli.update_add_cli_task(task, self.populate_template, self.connector, self.inputQueue)
            elif 'make_primary_creds' in task.task:
                make_primary_creds.update_make_primary_creds_task(task, self.populate_template, self.connector)
            elif 'add_snmpv2c' in task.task:
                add_snmpv2c.update_add_snmpv2c_task(task, self.populate_template, self.inputQueue)
            elif 'add_snmpv3' in task.task:
                add_snmpv3.update_add_snmpv3_task(task, self.populate_template, self.inputQueue)
            elif 'add_network_settings' in task.task:
                add_network_settings.update_add_network_settings_task(task, self.populate_template)
            elif 'add_wireless_profile' in task.task:
                add_wireless_profile.update_add_wireless_profile_task(task, self.populate_template, self.inputQueue)
            elif 'add_ent_wlan' in task.task:
                add_ent_wlan.update_add_ent_wlan_task(task, self.populate_template)
            elif 'add_fabric' in task.task:
                add_fabric.update_add_fabric_task(task, self.populate_template, self.connector)
            elif 'set_auth_template' in task.task:
                set_auth_template.update_set_auth_template_task(task, self.populate_template, self.connector)
                #TODO host onboarding of ports
                #TODO add ip pool to vn
                #TODO delete config just posted
                #TODO pull data from CRD, other documents
            else:
                print('task not found!')
                exit(1)

    def process_tasks(self):
        for task in self.inputQueue:
            # Call external module to get any prerequisites needed to populate template and PUT/POST
            if 'add_ip_pool' in task.task:
                add_ip_pool.update_ip_pool_task(task, self.populate_template)
            elif 'add_virtual_network' in task.task:
                add_virtual_network.update_add_virtual_network_task(task, self.populate_template)
            elif 'add_scalable_group_to_vn' in task.task:
                add_sg_to_vn.update_add_sg_to_vn_task(task, self.populate_template, self.connector)
            elif 'add_site_group' in task.task:
                add_site_group.update_add_sg_to_vn_task(task, self.populate_template, self.connector)
            elif 'add_servers' in task.task:
                add_servers.update_add_servers(task, self.populate_template, self.connector)
            elif 'add_ip_pool_to_vn' in task.task:
                add_ip_pool_to_vn.update_add_ip_pool_to_vn_task(task, self.populate_template, self.connector)
            elif 'add_ise_aaa' in task.task:
                add_ise_aaa.update_add_ise_aaa_task(task, self.populate_template)
            elif 'set_aaa_properties' in task.task:
                set_aaa_properties.update_set_aaa_properties_task(task, self.populate_template, self.connector)
            elif 'add_cli' in task.task:
                add_cli.update_add_cli_task(task, self.populate_template, self.connector, self.inputQueue)
            elif 'make_primary_creds' in task.task:
                make_primary_creds.update_make_primary_creds_task(task, self.populate_template, self.connector)
            elif 'add_snmpv2c' in task.task:
                add_snmpv2c.update_add_snmpv2c_task(task, self.populate_template, self.inputQueue)
            elif 'add_snmpv3' in task.task:
                add_snmpv3.update_add_snmpv3_task(task, self.populate_template, self.inputQueue)
            elif 'add_network_settings' in task.task:
                add_network_settings.update_add_network_settings_task(task, self.populate_template)
            elif 'add_wireless_profile' in task.task:
                add_wireless_profile.update_add_wireless_profile_task(task, self.populate_template, self.inputQueue)
            elif 'add_ent_wlan' in task.task:
                add_ent_wlan.update_add_ent_wlan_task(task, self.populate_template)
            elif 'add_fabric' in task.task:
                add_fabric.update_add_fabric_task(task, self.populate_template, self.connector)
            elif 'set_auth_template' in task.task:
                set_auth_template.update_set_auth_template_task(task, self.populate_template, self.connector)
            else:
                print('task not found!')
                exit(1)

            response = self.connector.parse_request(task)
            task.status = response.status_code
            response_json = json.loads(response.content)
            #   print(response_json)

            # If response ok (202) then add the response URL for follow up
            if task.status == 202:
                task.set_attribute(response_uri=BASE_URI[:-4] + response_json['response']['url'])
            # Otherwise, add the response message and response detail to task.reason for reporting purposes
            elif 'Unauthorized' in response_json['response']['message']:
                # Let get_service_token() handle this, and then rerun process_tasks()
                self.connector.get_service_token()
                self.process_tasks()
            elif 'BAD_REQUEST' in response_json['response']['errorCode']:
                task.status = 'Failed!'
                task.reason = response_json['response']['errorCode'] + response_json['response']['message']
            else:
                task.status = response_json['response']['message'] + response_json['response']['detail']

            self.check_job_status(task)

    def check_job_status_debug(self, task):
        while True:
            if task.status == 202:
                response = self.connector.get_request(task.response_uri)
                job_status = json.loads(response.content)
                print(job_status)
            time.sleep(1)

    def check_job_status(self, task):
        if task.status == 202:
            print("Checking job status, querying {}".format(task.response_uri))
            current_time = datetime.datetime.now()
            kill_time = current_time + datetime.timedelta(seconds=30)
            done = False
            while datetime.datetime.now() < kill_time and not done:
                response = self.connector.get_request(task.response_uri)
                job_status = json.loads(response.content)
                if 'endTime' not in job_status['response']:
                    print("Waiting for {} task to finish...".format(task.task[2]))
                    time.sleep(1)
                else:
                    task.complete_time = (job_status['response']['endTime'] - job_status['response'][
                        'startTime']) / 1000, 'sec'
                    if 'Failed' in job_status['response']['progress']:
                        task.status = job_status['response']['progress']
                        task.reason = job_status['response']['errorCode'], job_status['response']['failureReason']
                    else:
                        task.status = 'Success!'
                        task.reason = job_status['response']['serviceType']
                    print("....Task completed!")
                    done = True

    def print_report(self):
        if any('Success' in task.status for task in self.inputQueue):
            print(SEPARATOR)
            print("Successfull tasks:")
        successfull_tasks = (task for task in self.inputQueue if 'Success' in task.status)
        for task in successfull_tasks:
            task.print_task()

        if any('Failed' in task.status for task in self.inputQueue):
            print(SEPARATOR)
            print("Failed tasks:")
        failed_tasks = (task for task in self.inputQueue if 'Failed' in task.status)
        for task in failed_tasks:
            task.print_task()
        print(SEPARATOR)

    #
    # Print the contents of the inputQueue without purging values
    #
    def print_queued_tasks(self):
        print(SEPARATOR)
        print("DNAC Build Tasks Queued for Commit:")
        for i in range(0, len(self.inputQueue)):
            self.inputQueue[i].print_task()
        print(SEPARATOR)

    #
    # Populate the jinja2 templates in templates folder with values from spreadsheet
    #
    @staticmethod
    def populate_template(template, data):
        jinja_loader = jinja2.FileSystemLoader(CWD)
        jinja_env = jinja2.Environment(loader=jinja_loader, trim_blocks=True, lstrip_blocks=True)
        context_template = jinja_env.get_template(TEMPLATE_DIR + template)
        return context_template.render(data=data)

    #
    # Read the xls file defined in 'filename' and the workbook defined by 'wb_name,' create a dictionary for each row
    # using the headings values as keys. Return a list of dictionaries
    #
    def get_wb_data(self):
        for task in self.build_tasks:
            wb = load_workbook(self.filename)
            content = wb[task[2]]
            data = content.values
            headings = next(data)
            rows = list(data)
            for row in rows:
                self.inputQueue.append(
                    #DNACBuildTask(template=self.populate_template(task[3], dict(zip(headings, row))),
                    #              uri=URIs[task[2]], method=METHOD[task[2]], name=task[2]))
                    DNACBuildTask(template=dict(zip(headings, row)), task=task))

    #
    # Read the xls file defined in 'filename,' capture the headings and rows, create list of build_tasks where 1st
    # column under heading 'include' equals 'Yes'
    #
    @gate_keeper
    def get_build_tasks(self):
        try:
            wb = load_workbook(self.filename)
            build_tasks = wb['build_tasks']
            data = build_tasks.values
            headings = next(data)
            rows = list(data)
            [self.build_tasks.append(row) if 'Yes' in row[0] else '' for row in rows]
            print(SEPARATOR)
            print("Selected DNAC build tasks are as follows:")
            for r in self.build_tasks:
                print('+', '{:30}'.format(r[2]), r[1])
            print(SEPARATOR)
        except IOError:
            print("Invalid XLS file. Please check the XLS file in use")
            exit(1)


class DNACConnector(object):
    def __init__(self):
        self.dnac_token = ''
        self.creds = USER + ':' + PASSWORD
        self.b64cred = base64.b64encode(self.creds.encode('utf8'))
        self.headers = ''

    def update_headers(self):
        self.headers = {'Content-type': 'application/json', 'X-Auth-Token': self.dnac_token}

    def parse_request(self, task):
        if 'POST' in task.method:
            return self.post_request(task)
        elif 'PUT' in task.method:
            return self.put_request(task)
        else:
            return KeyError

    def post_request(self, task):
        return requests.post(task.uri, data=task.jinja, verify=False, headers=self.headers)

    def put_request(self, task):
        return requests.put(task.uri, data=task.jinja, verify=False, headers=self.headers)

    def get_request(self, uri):
        return requests.get(uri, verify=False, headers=self.headers)

    #
    # Try to get service token using credentials in dnac_settings.py file, if that fails call get_credentials()
    #
    def get_service_token(self):
        print(SEPARATOR)
        print("Updating DNA API Service Token")
        print(SEPARATOR)
        print(self.b64cred.decode('utf8'))
        # Request a new service token
        hyperURL = BASE_URI + "/system/v1/auth/token"
        response = requests.post(hyperURL, verify=False,
                                 headers={'Authorization': 'Basic {}'.format(self.b64cred.decode('utf8'))})
        print("DNAC IP Addr:", BASE_URI[8:-4], "Return Code: " + str(response.status_code))
        if response.status_code != 200:
            print("Login failed.  Please re-enter credentials:")
            self.get_credentials()
            self.get_service_token()
            return
        json_service_token = json.loads(response.content)
        self.dnac_token = json_service_token["Token"]
        print("DNA Token: " + self.dnac_token)
        self.update_headers()
        print(SEPARATOR)

    #
    # If credentials from dnac_settings prove to be incorrect, manually capture credentials using getpass()
    #
    def get_credentials(self):
        USER = input("DNA ADMIN USERNAME: ")
        PASSWORD = getpass.getpass("DNA ADMIN PASSWORD: ")
        creds = USER + ':' + PASSWORD
        self.b64cred = base64.b64encode(creds.encode('utf8'))


#
# Its nice to know things so lets log stuff
#
class Logger(object):
    def __init__(self):
        self.log = ''
